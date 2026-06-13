"""Excel (.xlsx) dan mahsulotlarni import — ustun xaritasi va upsert."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from typing import Any

from django.db import transaction
from openpyxl import load_workbook

from .barcode_lookup import (
    find_product_by_barcode,
    normalize_barcode,
    sync_product_barcodes,
)
from .models import (
    Brand,
    Category,
    PriceList,
    Product,
    ProductBarcode,
    ProductPrice,
    Supplier,
    UnitOfMeasure,
)


def _normalize_name(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def _normalize_barcode(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float):
        if value != value:
            return ""
        if value == int(value):
            return str(int(value))
        return str(int(value)) if float(value).is_integer() else str(value).strip()
    s = str(value).strip()
    if not s or s.lower() in ("none", "nan"):
        return ""
    if s.endswith(".0") and s[:-2].replace("-", "").isdigit():
        s = s[:-2]
    return normalize_barcode(s)


def _parse_price(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return None
    try:
        if isinstance(value, str):
            cleaned = value.replace(" ", "").replace(",", ".")
            if not cleaned:
                return None
            return Decimal(cleaned).quantize(Decimal("0.01"))
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return None


def _parse_quantity(value: Any) -> Decimal | None:
    if value is None or value == "":
        return None
    try:
        if isinstance(value, str):
            cleaned = value.replace(" ", "").replace(",", ".")
            if not cleaned:
                return None
            return Decimal(cleaned)
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _parse_barcodes_from_cell(value: Any) -> list[str]:
    """Bitta katak: 2+ shtrix kod (Enter, vergul, ustun) yoki bitta kod."""
    if value is None:
        return []
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        code = _normalize_barcode(value)
        return [code] if code else []

    raw = str(value).strip()
    if not raw:
        return []

    raw = (
        raw.replace("\r\n", "\n")
        .replace("\r", "\n")
        .replace("\x0b", "\n")
        .replace("\u2028", "\n")
        .replace("\u2029", "\n")
    )

    parts = re.split(r"[,;|\n\t]+|\s{2,}", raw)
    codes: list[str] = []
    seen: set[str] = set()
    for part in parts:
        code = _normalize_barcode(part)
        if code and code not in seen:
            seen.add(code)
            codes.append(code)
    if codes:
        return codes

    single = _normalize_barcode(value)
    return [single] if single else []


def _get_cell(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx < 0 or idx >= len(row):
        return None
    return row[idx]


def _load_rows(uploaded_file) -> list[tuple[Any, ...]]:
    wb = load_workbook(uploaded_file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    return rows


def get_import_target_options(tenant) -> list[dict[str, str]]:
    """Backendda mavjud barcha qabul qiluvchi maydonlar."""
    options: list[dict[str, str]] = [
        {"key": "skip", "label": "— O'tkazib yuborish —"},
        {"key": "name", "label": "Mahsulot nomi (majburiy)"},
        {"key": "full_name", "label": "To'liq nom"},
        {"key": "barcode", "label": "Shtrix kod / QR"},
        {"key": "cost_price", "label": "Kirish narxi"},
        {"key": "sku", "label": "Artikul (SKU)"},
        {"key": "unit", "label": "O'lchov birligi"},
        {"key": "quantity", "label": "Ombordagi miqdor"},
        {"key": "category", "label": "Bo'lim (kategoriya)"},
        {"key": "brand", "label": "Brend"},
        {"key": "supplier", "label": "Ta'minotchi"},
    ]

    price_lists = PriceList.objects.filter(tenant=tenant, is_active=True).order_by(
        "sort_order", "name"
    )
    has_selling = False
    for pl in price_lists:
        if pl.is_selling:
            options.append({"key": "price", "label": f"Sotuv narxi ({pl.name})"})
            has_selling = True
        else:
            options.append(
                {"key": f"price_list:{pl.id}", "label": f"Narx: {pl.name}"}
            )
    if not has_selling:
        options.insert(5, {"key": "price", "label": "Sotuv narxi (kassa)"})

    return options


def preview_excel_columns(uploaded_file, tenant=None) -> dict:
    rows = _load_rows(uploaded_file)
    if not rows:
        result = {"columns": [], "sample_rows": [], "suggested_targets": {}}
        if tenant:
            result["target_options"] = get_import_target_options(tenant)
        return result

    header = rows[0]
    columns = []
    for idx, cell in enumerate(header):
        label = str(cell).strip() if cell is not None else ""
        columns.append({"index": idx, "label": label or f"Ustun {idx + 1}"})

    sample_rows = []
    for row in rows[1:4]:
        sample_rows.append([_cell_preview(c) for c in row])

    result = {
        "columns": columns,
        "sample_rows": sample_rows,
        "suggested_targets": _suggest_targets(header, tenant),
    }
    if tenant:
        result["target_options"] = get_import_target_options(tenant)
    return result


def _cell_preview(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value)[:80]


def _suggest_targets(header: tuple[Any, ...], tenant=None) -> dict[str, str]:
    """Excel ustun indeksi (string) → maydon kaliti."""
    suggestions: dict[str, str] = {}
    name_assigned = False
    price_lists: list[PriceList] = []
    if tenant:
        price_lists = list(
            PriceList.objects.filter(tenant=tenant, is_active=True).order_by(
                "sort_order", "name"
            )
        )

    for idx, cell in enumerate(header):
        if cell is None:
            continue
        label = str(cell).lower().strip()
        if not label:
            continue

        matched_price_list = False
        for pl in price_lists:
            pl_name = pl.name.lower().strip()
            if pl_name and (pl_name == label or pl_name in label):
                suggestions[str(idx)] = (
                    "price" if pl.is_selling else f"price_list:{pl.id}"
                )
                matched_price_list = True
                break
        if matched_price_list:
            continue

        if label in ("optom", "chakana", "wholesale", "retail"):
            for pl in price_lists:
                if pl.name.lower().strip() == label:
                    suggestions[str(idx)] = (
                        "price" if pl.is_selling else f"price_list:{pl.id}"
                    )
                    matched_price_list = True
                    break
        if matched_price_list:
            continue

        if not name_assigned and (
            "наименование" in label or label.startswith("номи") or label == "name"
        ) and "полное" not in label and "полн" not in label:
            suggestions[str(idx)] = "name"
            name_assigned = True
        elif "полное" in label and "наимен" in label:
            suggestions[str(idx)] = "full_name"
        elif any(k in label for k in ("штрих", "shtrix", "barcode", "qr", "bar kod")):
            suggestions[str(idx)] = "barcode"
        elif any(k in label for k in ("цена", "narx", "price", "продаж", "sotish")):
            suggestions[str(idx)] = "price"
        elif any(k in label for k in ("себест", "kirish", "cost", "закуп")):
            suggestions[str(idx)] = "cost_price"
        elif any(k in label for k in ("артикул", "sku", "artikul")):
            suggestions[str(idx)] = "sku"
        elif any(k in label for k in ("мера", "едини", "unit", "o'lchov", "olchov")):
            suggestions[str(idx)] = "unit"
        elif any(k in label for k in ("колич", "miqdor", "quantity", "qoldiq", "остат")):
            suggestions[str(idx)] = "quantity"
        elif any(
            k in label
            for k in ("катег", "bo'lim", "bolim", "тип прод", "category", "раздел")
        ):
            suggestions[str(idx)] = "category"
        elif any(k in label for k in ("brend", "brand", "бренд", "производ")):
            suggestions[str(idx)] = "brand"
        elif any(k in label for k in ("ta'minot", "tamino", "поставщ", "supplier")):
            suggestions[str(idx)] = "supplier"

    return suggestions


def _parse_column_targets(raw: Any) -> dict[str, Any]:
    """Frontend: { "0": "name", "2": "unit", "5": "barcode" }"""
    if isinstance(raw, str):
        try:
            raw = json.loads(raw)
        except json.JSONDecodeError:
            return {}
    if not isinstance(raw, dict):
        return {}

    field_map: dict[str, Any] = {
        "name": None,
        "full_name": None,
        "unit": None,
        "sku": None,
        "price": None,
        "cost_price": None,
        "quantity": None,
        "category": None,
        "brand": None,
        "supplier": None,
        "barcode": [],
        "price_lists": {},
    }

    for idx_str, target in raw.items():
        try:
            idx = int(idx_str)
        except (TypeError, ValueError):
            continue
        t = (target or "skip").strip().lower()
        if t == "skip" or not t:
            continue
        if t == "barcode":
            if idx not in field_map["barcode"]:
                field_map["barcode"].append(idx)
        elif t.startswith("price_list:"):
            pl_id = target.split(":", 1)[1]
            field_map["price_lists"][pl_id] = idx
        elif t in field_map:
            field_map[t] = idx

    return field_map


def _detect_legacy_columns(header: tuple[Any, ...]) -> dict[str, Any]:
    name_keys = ("наименование", "номи", "mahsulot", "товар", "name", "product", "nomi")
    barcode_keys = ("штрих", "shtrix", "barcode", "bar kod", "qr")
    price_keys = ("цена", "narx", "price", "продаж", "sum", "sotish")

    name_idx: int | None = None
    price_idx: int | None = None
    barcode_indices: list[int] = []

    for idx, cell in enumerate(header):
        if cell is None:
            continue
        label = str(cell).lower().strip()
        if not label:
            continue
        if name_idx is None and any(k in label for k in name_keys):
            name_idx = idx
        elif price_idx is None and any(k in label for k in price_keys):
            price_idx = idx
        elif any(k in label for k in barcode_keys):
            barcode_indices.append(idx)

    if name_idx is not None:
        return {
            "name": name_idx,
            "price": price_idx,
            "cost_price": None,
            "unit": None,
            "sku": None,
            "quantity": None,
            "category": None,
            "full_name": None,
            "barcode": sorted(set(barcode_indices)),
        }
    return {}


def _collect_row_barcodes(row: tuple[Any, ...], barcode_indices: list[int]) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    for idx in barcode_indices:
        for code in _parse_barcodes_from_cell(_get_cell(row, idx)):
            if code not in seen:
                seen.add(code)
                codes.append(code)
    return codes


def _resolve_unit(tenant, unit_text: str) -> tuple[str, UnitOfMeasure | None]:
    text = _normalize_name(unit_text)
    if not text or text.lower() in ("не указано", "—", "-"):
        return "dona", None

    weighable = any(
        w in text.lower() for w in ("kg", "кг", "gr", "гр", "gram", "litr", "литр")
    )
    unit_obj, _ = UnitOfMeasure.objects.get_or_create(
        tenant=tenant,
        name=text[:50],
        defaults={"is_weighable": weighable},
    )
    return text[:20], unit_obj


def _resolve_category(tenant, cat_text: str) -> Category | None:
    text = _normalize_name(cat_text)
    if not text or text.lower() in ("не указано", "—", "-", "barchasi"):
        return None
    cat, _ = Category.objects.get_or_create(
        tenant=tenant,
        name=text[:200],
        defaults={"sort_order": 0},
    )
    return cat


def _resolve_brand(tenant, brand_text: str) -> Brand | None:
    text = _normalize_name(brand_text)
    if not text or text.lower() in ("не указано", "—", "-"):
        return None
    brand, _ = Brand.objects.get_or_create(
        tenant=tenant,
        name=text[:120],
        defaults={"is_active": True},
    )
    return brand


def _resolve_supplier(tenant, supplier_text: str) -> Supplier | None:
    text = _normalize_name(supplier_text)
    if not text or text.lower() in ("не указано", "—", "-"):
        return None
    supplier, _ = Supplier.objects.get_or_create(
        tenant=tenant,
        name=text[:200],
    )
    return supplier


def _sync_import_list_prices(tenant, product: Product, list_prices: dict[str, Decimal]):
    if not list_prices:
        return
    active_ids = {
        str(x)
        for x in PriceList.objects.filter(
            tenant=tenant, is_active=True, is_selling=False
        ).values_list("id", flat=True)
    }
    for pl_id, amount in list_prices.items():
        if pl_id not in active_ids:
            continue
        ProductPrice.objects.update_or_create(
            tenant=tenant,
            product=product,
            price_list_id=pl_id,
            defaults={"price": amount},
        )


def _find_existing_product(
    tenant, name: str, barcodes: list[str], sku: str
) -> Product | None:
    for code in barcodes:
        product = find_product_by_barcode(tenant, code)
        if product:
            return product

    if sku:
        product = Product.objects.filter(tenant=tenant, sku__iexact=sku).first()
        if product:
            return product

    normalized_name = _normalize_name(name)
    if normalized_name:
        return Product.objects.filter(tenant=tenant, name__iexact=normalized_name).first()

    return None


def _merge_barcodes(product: Product, import_barcodes: list[str], tenant) -> list[str]:
    existing = list(product.barcodes.values_list("code", flat=True))
    if not existing and product.barcode:
        existing = [product.barcode]

    merged: list[str] = []
    seen: set[str] = set()
    for code in existing + import_barcodes:
        c = normalize_barcode(code)
        if not c or c in seen:
            continue
        taken = (
            ProductBarcode.objects.filter(tenant=tenant, code=c)
            .exclude(product=product)
            .exists()
        )
        if taken:
            continue
        seen.add(c)
        merged.append(c)
    return merged


def _process_product(
    tenant,
    *,
    name: str,
    barcodes: list[str],
    price: Decimal | None,
    cost_price: Decimal | None = None,
    unit_text: str = "",
    sku: str = "",
    quantity: Decimal | None = None,
    category_name: str = "",
    brand_name: str = "",
    supplier_name: str = "",
    list_prices: dict[str, Decimal] | None = None,
    full_name: str = "",
) -> tuple[str, int]:
    display_name = _normalize_name(name)
    if full_name and _normalize_name(full_name) != display_name:
        display_name = _normalize_name(full_name) or display_name

    if not display_name:
        raise ValueError("Mahsulot nomi bo'sh")

    sku_clean = (sku or "").strip()[:64]

    before_codes = set(
        ProductBarcode.objects.filter(tenant=tenant).values_list("code", flat=True)
    )

    product = _find_existing_product(tenant, display_name, barcodes, sku_clean)
    is_new = product is None

    unit_str, unit_ref = _resolve_unit(tenant, unit_text) if unit_text else ("dona", None)
    category = _resolve_category(tenant, category_name) if category_name else None
    brand = _resolve_brand(tenant, brand_name) if brand_name else None
    supplier = _resolve_supplier(tenant, supplier_name) if supplier_name else None
    list_prices = list_prices or {}

    if is_new:
        create_price = price if price is not None else Decimal("0")
        if create_price == 0 and list_prices:
            create_price = next(iter(list_prices.values()))
        product = Product.objects.create(
            tenant=tenant,
            name=display_name,
            price=create_price,
            cost_price=cost_price or Decimal("0"),
            sku=sku_clean,
            unit=unit_str,
            unit_ref=unit_ref,
            category=category,
            brand=brand,
            supplier=supplier,
            quantity=quantity or Decimal("0"),
            barcode=barcodes[0] if barcodes else "",
        )
        action = "created"
        final_barcodes = barcodes
    else:
        action = "updated"
        final_barcodes = (
            _merge_barcodes(product, barcodes, tenant) if barcodes else []
        )

        update_fields: list[str] = []

        if price is not None and product.price != price:
            product.price = price
            update_fields.append("price")
        if cost_price is not None and product.cost_price != cost_price:
            product.cost_price = cost_price
            update_fields.append("cost_price")
        if sku_clean and product.sku != sku_clean:
            product.sku = sku_clean
            update_fields.append("sku")
        if unit_text and (product.unit != unit_str or product.unit_ref_id != (
            unit_ref.id if unit_ref else None
        )):
            product.unit = unit_str
            product.unit_ref = unit_ref
            update_fields.extend(["unit", "unit_ref"])
        if quantity is not None and product.quantity != quantity:
            product.quantity = quantity
            update_fields.append("quantity")
        if category and product.category_id != category.id:
            product.category = category
            update_fields.append("category")
        if brand and product.brand_id != brand.id:
            product.brand = brand
            update_fields.append("brand")
        if supplier and product.supplier_id != supplier.id:
            product.supplier = supplier
            update_fields.append("supplier")
        if display_name and product.name != display_name:
            product.name = display_name
            update_fields.append("name")

        if update_fields:
            update_fields.append("updated_at")
            product.save(update_fields=list(dict.fromkeys(update_fields)))

    if list_prices:
        _sync_import_list_prices(tenant, product, list_prices)
        if price is None and list_prices:
            primary = next(iter(list_prices.values()))
            if product.price != primary:
                product.price = primary
                product.save(update_fields=["price", "updated_at"])

    if final_barcodes:
        sync_product_barcodes(product, final_barcodes)
    elif is_new and not product.barcode:
        product.save(update_fields=["barcode", "updated_at"])

    after_codes = set(
        ProductBarcode.objects.filter(tenant=tenant, product=product).values_list(
            "code", flat=True
        )
    )
    added = len(after_codes - before_codes)
    return action, added


def import_products_from_excel(
    tenant, uploaded_file, column_targets: dict | None = None
) -> dict[str, int | list]:
    rows = _load_rows(uploaded_file)

    empty_stats: dict[str, int | list] = {
        "created_products": 0,
        "updated_products": 0,
        "barcodes_added": 0,
        "rows_skipped": 0,
        "rows_processed": 0,
        "total_rows": 0,
        "row_errors": [],
    }

    if not rows:
        return empty_stats

    if column_targets:
        cols = _parse_column_targets(column_targets)
    else:
        cols = _detect_legacy_columns(rows[0])
        if cols.get("name") is None:
            cols = _parse_column_targets(_suggest_targets(rows[0], tenant))

    name_idx = cols.get("name")
    if name_idx is None:
        empty_stats["row_errors"] = ["«Mahsulot nomi» ustuni tanlanmagan."]
        return empty_stats

    data_rows = rows[1:]

    grouped: dict[str, dict[str, Any]] = defaultdict(
        lambda: {
            "barcodes": [],
            "price": None,
            "cost_price": None,
            "unit": "",
            "sku": "",
            "quantity": None,
            "category": "",
            "brand": "",
            "supplier": "",
            "list_prices": {},
            "full_name": "",
            "rows": [],
        }
    )

    rows_skipped = 0
    last_product_key: str | None = None
    barcode_cols = cols.get("barcode") or []

    for row_num, row in enumerate(data_rows, start=2):
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        row_barcodes = _collect_row_barcodes(row, barcode_cols)
        name = _normalize_name(_get_cell(row, name_idx))

        if name:
            key = name.casefold()
            last_product_key = key
        elif last_product_key and row_barcodes:
            key = last_product_key
            name = grouped[last_product_key]["name"]
        elif row_barcodes:
            key = None
            for code in row_barcodes:
                existing = find_product_by_barcode(tenant, code)
                if existing:
                    key = existing.name.casefold()
                    name = existing.name
                    last_product_key = key
                    break
            if not key:
                rows_skipped += 1
                continue
        else:
            rows_skipped += 1
            continue

        price = _parse_price(_get_cell(row, cols.get("price")))
        cost_price = _parse_price(_get_cell(row, cols.get("cost_price")))
        quantity = _parse_quantity(_get_cell(row, cols.get("quantity")))
        unit_val = _normalize_name(_get_cell(row, cols.get("unit"))) or ""
        sku_val = str(_get_cell(row, cols.get("sku")) or "").strip()
        cat_val = _normalize_name(_get_cell(row, cols.get("category"))) or ""
        brand_val = _normalize_name(_get_cell(row, cols.get("brand"))) or ""
        supplier_val = _normalize_name(_get_cell(row, cols.get("supplier"))) or ""
        full_name = _normalize_name(_get_cell(row, cols.get("full_name"))) or ""

        entry = grouped[key]
        entry["name"] = name
        entry["full_name"] = full_name or entry.get("full_name") or ""
        if price is not None:
            entry["price"] = price
        if cost_price is not None:
            entry["cost_price"] = cost_price
        if quantity is not None:
            entry["quantity"] = quantity
        if unit_val:
            entry["unit"] = unit_val
        if sku_val:
            entry["sku"] = sku_val
        if cat_val:
            entry["category"] = cat_val
        if brand_val:
            entry["brand"] = brand_val
        if supplier_val:
            entry["supplier"] = supplier_val
        for pl_id, col_idx in (cols.get("price_lists") or {}).items():
            pl_price = _parse_price(_get_cell(row, col_idx))
            if pl_price is not None:
                entry["list_prices"][pl_id] = pl_price
        entry["rows"].append(row_num)

        seen = set(entry["barcodes"])
        for code in row_barcodes:
            if code not in seen:
                seen.add(code)
                entry["barcodes"].append(code)

    created_products = 0
    updated_products = 0
    barcodes_added = 0
    rows_processed = 0
    row_errors: list[str] = []

    for entry in grouped.values():
        name = entry["name"]
        barcodes: list[str] = entry["barcodes"]

        try:
            with transaction.atomic():
                action, added = _process_product(
                    tenant,
                    name=name,
                    barcodes=barcodes,
                    price=entry.get("price"),
                    cost_price=entry.get("cost_price"),
                    unit_text=entry.get("unit") or "",
                    sku=entry.get("sku") or "",
                    quantity=entry.get("quantity"),
                    category_name=entry.get("category") or "",
                    brand_name=entry.get("brand") or "",
                    supplier_name=entry.get("supplier") or "",
                    list_prices=entry.get("list_prices") or {},
                    full_name=entry.get("full_name") or "",
                )
            rows_processed += len(entry["rows"])
            if action == "created":
                created_products += 1
            else:
                updated_products += 1
            barcodes_added += added
        except Exception as exc:
            rows_skipped += len(entry["rows"])
            row_nums = ", ".join(str(n) for n in entry["rows"][:3])
            if len(row_errors) < 20:
                row_errors.append(f"Qator(lar) {row_nums}: {exc}")

    return {
        "created_products": created_products,
        "updated_products": updated_products,
        "barcodes_added": barcodes_added,
        "rows_skipped": rows_skipped,
        "rows_processed": rows_processed,
        "total_rows": len(data_rows),
        "row_errors": row_errors,
    }
