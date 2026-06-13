"""Mahsulot import shabloni (.xlsx)."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .models import PriceList


def _price_headers(tenant) -> list[str]:
    if not tenant:
        return ["Цена для продажи SUM", "optom SUM"]
    headers: list[str] = []
    for pl in PriceList.objects.filter(tenant=tenant, is_active=True).order_by(
        "sort_order", "name"
    ):
        headers.append(f"{pl.name} SUM")
    if not headers:
        headers = ["Цена для продажи SUM", "optom SUM"]
    return headers


def build_import_template(tenant=None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Mahsulotlar"

    price_headers = _price_headers(tenant)
    headers = [
        "Наименование товара",
        "Штрих-код",
        "Штрих код 2",
        "Штрих код 3",
        *price_headers,
    ]
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF")
    wrap = Alignment(wrap_text=True, vertical="top")

    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=title)
        cell.fill = header_fill
        cell.font = header_font

    price_col_start = 5
    selling_col = price_col_start
    wholesale_col = price_col_start + 1 if len(price_headers) > 1 else None

    # 1-usul: bitta katakda bir nechta shtrix kod (Enter bilan)
    ws.cell(row=2, column=1, value="ACTIMEL 270 gr")
    ws.cell(row=2, column=2, value="4600605034446\n4600605034453")
    ws.cell(row=2, column=2).alignment = wrap
    ws.cell(row=2, column=selling_col, value=15300)
    if wholesale_col:
        ws.cell(row=2, column=wholesale_col, value=14200)

    # 2-usul: bir nechta ustunda
    ws.cell(row=3, column=1, value="MARI MALAKO 1000 ml")
    ws.cell(row=3, column=2, value="4640019490100")
    ws.cell(row=3, column=3, value="4640019491572")
    ws.cell(row=3, column=4, value="8600135890520")
    ws.cell(row=3, column=selling_col, value=28500)
    if wholesale_col:
        ws.cell(row=3, column=wholesale_col, value=26000)

    samples = [
        ("KAIJA SHPROTI 240 gr", "4751007730624", "", 29000, 27500),
        ("KAIJA SHPROTI 240 gr", "4780036863706", "", 29000, 27500),
        ("NON GAMBURGER", "23", "", 24000, 22000),
    ]
    for row_idx, (name, bc1, bc2, sell_price, wholesale_price) in enumerate(samples, start=4):
        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=bc1)
        ws.cell(row=row_idx, column=3, value=bc2)
        ws.cell(row=row_idx, column=selling_col, value=sell_price)
        if wholesale_col:
            ws.cell(row=row_idx, column=wholesale_col, value=wholesale_price)

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    for offset in range(len(price_headers)):
        col_letter = chr(ord("E") + offset)
        ws.column_dimensions[col_letter].width = 22

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
